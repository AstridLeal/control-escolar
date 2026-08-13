# Rutas relacionadas con reportes y exportación de datos
from flask import Blueprint, request, jsonify, send_file, session
from backend.routes.auth import require_auth
from backend.models.pago import Pago
from backend.models.asistencia import Asistencia
from backend.models.estudiante import Estudiante
from backend.models.calificacion import Calificacion
from backend.utils.database import execute_query
from io import BytesIO
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib.pagesizes import A4, landscape
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.units import inch
from datetime import datetime

# Blueprint para las rutas de reportes
reportes_bp = Blueprint('reportes', __name__)

# Funciones auxiliares para generar estilos y respuestas de archivos
def _header_style():
    return Font(bold=True, color='FFFFFF'), PatternFill('solid', fgColor='2563EB')

# Función para generar una tabla en PDF con estilo
def _pdf_table(data, col_widths=None):
    table = Table(data, colWidths=col_widths)
    table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2563EB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 8),
        ('GRID', (0, 0), (-1, -1), 0.5, colors.grey),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F1F5F9')]),
    ]))
    return table

# Función para generar una respuesta de archivo Excel
def _excel_response(wb, filename):
    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    return send_file(
        buffer,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )

# Función para generar una respuesta de archivo PDF
def _pdf_response(elements, filename, landscape_mode=False):
    buffer = BytesIO()
    pagesize = landscape(A4) if landscape_mode else A4
    doc = SimpleDocTemplate(buffer, pagesize=pagesize, leftMargin=0.5*inch, rightMargin=0.5*inch)
    doc.build(elements)
    buffer.seek(0)
    return send_file(buffer, mimetype='application/pdf', as_attachment=True, download_name=filename)

# Rutas para reportes y exportación de datos
@reportes_bp.route('/dashboard', methods=['GET'])
def dashboard_general():
    ok, err, code = require_auth(['admin', 'secretaria'])()
    if not ok:
        return err, code
    id_periodo = request.args.get('id_periodo', type=int)

    total_estudiantes = execute_query(
        "SELECT COUNT(*) AS c FROM estudiante WHERE activo = TRUE", fetch_one=True
    )['c']
    total_docentes = execute_query(
        "SELECT COUNT(*) AS c FROM docente WHERE activo = TRUE", fetch_one=True
    )['c']
    total_matriculas = execute_query(
        "SELECT COUNT(*) AS c FROM estudiante WHERE activo = TRUE AND id_seccion IS NOT NULL",
        fetch_one=True
    )['c']
    pagos_dash = Pago.dashboard_morosos(id_periodo)

    return jsonify({
        'total_estudiantes': total_estudiantes,
        'total_docentes': total_docentes,
        'total_matriculas': total_matriculas,
        'pagos': pagos_dash
    })

# Rutas para exportar reportes en PDF o Excel
@reportes_bp.route('/exportar/pagos', methods=['GET'])
def exportar_pagos():
    ok, err, code = require_auth(['admin', 'secretaria'])()
    if not ok:
        return err, code
    formato = request.args.get('formato', 'excel')
    pagos = Pago.listar(
        id_periodo=request.args.get('id_periodo', type=int),
        estado=request.args.get('estado'),
        id_estudiante=request.args.get('id_estudiante', type=int),
        mes=request.args.get('mes'),
        anio=request.args.get('anio', type=int)
    )

    if formato == 'pdf':
        styles = getSampleStyleSheet()
        elements = [
            Paragraph("Reporte de Pagos - Control Escolar", styles['Title']),
            Spacer(1, 12),
            Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']),
            Spacer(1, 20)
        ]
        data = [['Estudiante', 'Cédula', 'Concepto', 'Monto', 'Mes/Año', 'Estado', 'Fecha']]
        for p in pagos:
            data.append([
                p.get('estudiante_nombre', ''), p.get('cedula', ''),
                p.get('concepto', ''), f"${float(p.get('monto', 0)):.2f}",
                f"{p.get('mes') or '-'}/{p.get('anio') or '-'}",
                p.get('estado', ''), str(p.get('fecha_pago') or '-')
            ])
        elements.append(_pdf_table(data))
        return _pdf_response(elements, 'reporte_pagos.pdf')

    wb = Workbook()
    ws = wb.active
    ws.title = "Pagos"
    headers = ['ID', 'Estudiante', 'Cédula', 'Concepto', 'Monto', 'Mes', 'Año', 'Estado', 'Fecha Pago']
    ws.append(headers)
    hf, hfill = _header_style()
    for col in range(1, len(headers) + 1):
        ws.cell(1, col).font = hf
        ws.cell(1, col).fill = hfill
    for p in pagos:
        ws.append([
            p.get('id_pago'), p.get('estudiante_nombre'), p.get('cedula'),
            p.get('concepto'), float(p.get('monto', 0)), p.get('mes'), p.get('anio'),
            p.get('estado'), str(p.get('fecha_pago') or '')
        ])
    return _excel_response(wb, 'reporte_pagos.xlsx')

# Rutas para exportar reportes de matrículas, asistencias y calificaciones
@reportes_bp.route('/exportar/matriculas', methods=['GET'])
def exportar_matriculas():
    ok, err, code = require_auth(['admin', 'secretaria'])()
    if not ok:
        return err, code
    formato = request.args.get('formato', 'excel')
    estudiantes = [e for e in Estudiante.listar() if e.get('id_seccion')]
    id_grado = request.args.get('id_grado', type=int)
    id_seccion = request.args.get('id_seccion', type=int)
    if id_grado:
        estudiantes = [e for e in estudiantes if e.get('id_grado') == id_grado]
    if id_seccion:
        estudiantes = [e for e in estudiantes if e.get('id_seccion') == id_seccion]

    if formato == 'pdf':
        styles = getSampleStyleSheet()
        elements = [
            Paragraph("Reporte de Matrículas", styles['Title']),
            Spacer(1, 12),
            Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']),
            Spacer(1, 20)
        ]
        data = [['Estudiante', 'Cédula', 'Grado', 'Sección', 'Fecha']]
        for e in estudiantes:
            data.append([
                f"{e.get('nombres', '')} {e.get('apellidos', '')}",
                e.get('cedula', ''), e.get('grado_nombre', ''),
                e.get('seccion_nombre', ''), str(e.get('fecha_inscripcion') or '')
            ])
        elements.append(_pdf_table(data))
        return _pdf_response(elements, 'reporte_matriculas.pdf')

    wb = Workbook()
    ws = wb.active
    ws.title = "Matrículas"
    headers = ['Estudiante', 'Cédula', 'Grado', 'Sección', 'Periodo', 'Fecha Inscripción']
    ws.append(headers)
    hf, hfill = _header_style()
    for col in range(1, len(headers) + 1):
        ws.cell(1, col).font = hf
        ws.cell(1, col).fill = hfill
    for e in estudiantes:
        ws.append([
            f"{e.get('nombres', '')} {e.get('apellidos', '')}",
            e.get('cedula'), e.get('grado_nombre'), e.get('seccion_nombre'),
            e.get('periodo_nombre'), str(e.get('fecha_inscripcion') or '')
        ])
    return _excel_response(wb, 'reporte_matriculas.xlsx')

# Rutas para exportar reportes de asistencias y calificaciones
@reportes_bp.route('/exportar/asistencias', methods=['GET'])
def exportar_asistencias():
    ok, err, code = require_auth(['admin', 'secretaria', 'docente', 'estudiante'])()
    if not ok:
        return err, code

    # Alumno solo puede exportar las suyas
    id_estudiante = request.args.get('id_estudiante', type=int)
    if session.get('rol') == 'estudiante':
        id_estudiante = session.get('id_estudiante')
        if not id_estudiante:
            return jsonify({'error': 'Sin estudiante vinculado'}), 400

    formato = request.args.get('formato', 'excel')
    asistencias = Asistencia.listar(
        fecha=request.args.get('fecha'),
        id_seccion=request.args.get('id_seccion', type=int),
        id_grado=request.args.get('id_grado', type=int),
        id_estudiante=id_estudiante,
        fecha_desde=request.args.get('fecha_desde'),
        fecha_hasta=request.args.get('fecha_hasta')
    )

    if formato == 'pdf':
        styles = getSampleStyleSheet()
        elements = [
            Paragraph("Historial de Asistencias", styles['Title']),
            Spacer(1, 8),
            Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']),
            Spacer(1, 16)
        ]
        data = [['Fecha', 'Estudiante', 'Cédula', 'Grado', 'Sección', 'Estado', 'Obs.']]
        for a in asistencias:
            data.append([
                str(a.get('fecha') or ''), a.get('estudiante_nombre', ''),
                a.get('cedula', ''), a.get('grado_nombre', '') or '',
                a.get('seccion_nombre', '') or '', a.get('estado', ''),
                a.get('observacion') or ''
            ])
        elements.append(_pdf_table(data))
        return _pdf_response(elements, 'reporte_asistencias.pdf', landscape_mode=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Asistencias"
    headers = ['Fecha', 'Estudiante', 'Cédula', 'Grado', 'Sección', 'Estado', 'Observación']
    ws.append(headers)
    hf, hfill = _header_style()
    for col in range(1, len(headers) + 1):
        ws.cell(1, col).font = hf
        ws.cell(1, col).fill = hfill
    for a in asistencias:
        ws.append([
            str(a.get('fecha') or ''), a.get('estudiante_nombre'), a.get('cedula'),
            a.get('grado_nombre'), a.get('seccion_nombre'), a.get('estado'),
            a.get('observacion')
        ])
    return _excel_response(wb, 'reporte_asistencias.xlsx')

# Rutas para exportar reportes de calificaciones
@reportes_bp.route('/exportar/calificaciones', methods=['GET'])
def exportar_calificaciones():
    ok, err, code = require_auth(['admin', 'secretaria', 'docente', 'estudiante'])()
    if not ok:
        return err, code

    id_estudiante = request.args.get('id_estudiante', type=int)
    if session.get('rol') == 'estudiante':
        id_estudiante = session.get('id_estudiante')

    formato = request.args.get('formato', 'excel')
    id_materia = request.args.get('id_materia', type=int)
    id_periodo = request.args.get('id_periodo', type=int)
    id_seccion = request.args.get('id_seccion', type=int)

    # Si piden acta (materia + sección + periodo)
    if id_materia and id_seccion and id_periodo and not id_estudiante:
        rows = Calificacion.acta(id_materia, id_seccion, id_periodo)
        titulo = "Acta de Calificaciones"
    else:
        rows = Calificacion.listar(
            id_estudiante=id_estudiante,
            id_materia=id_materia,
            id_periodo=id_periodo,
            id_seccion=id_seccion
        )
        titulo = "Reporte de Calificaciones"

    if formato == 'pdf':
        styles = getSampleStyleSheet()
        elements = [
            Paragraph(titulo, styles['Title']),
            Spacer(1, 8),
            Paragraph(f"Generado: {datetime.now().strftime('%d/%m/%Y %H:%M')}", styles['Normal']),
            Spacer(1, 16)
        ]
        if id_materia and id_seccion and id_periodo and not id_estudiante:
            data = [['Cédula', 'Estudiante', 'N1', 'N2', 'N3', 'Examen', 'Promedio']]
            for r in rows:
                data.append([
                    r.get('cedula', ''),
                    f"{r.get('nombres', '')} {r.get('apellidos', '')}",
                    str(r.get('nota1') if r.get('nota1') is not None else '-'),
                    str(r.get('nota2') if r.get('nota2') is not None else '-'),
                    str(r.get('nota3') if r.get('nota3') is not None else '-'),
                    str(r.get('examen_final') if r.get('examen_final') is not None else '-'),
                    str(r.get('nota_final') if r.get('nota_final') is not None else '-'),
                ])
        else:
            data = [['Estudiante', 'Cédula', 'Materia', 'Periodo', 'N1', 'N2', 'N3', 'Examen', 'Promedio']]
            for r in rows:
                data.append([
                    r.get('estudiante_nombre', '') or f"{r.get('nombres', '')} {r.get('apellidos', '')}",
                    r.get('cedula', ''),
                    r.get('materia_nombre', ''),
                    r.get('periodo_nombre', ''),
                    str(r.get('nota1') if r.get('nota1') is not None else '-'),
                    str(r.get('nota2') if r.get('nota2') is not None else '-'),
                    str(r.get('nota3') if r.get('nota3') is not None else '-'),
                    str(r.get('examen_final') if r.get('examen_final') is not None else '-'),
                    str(r.get('nota_final') if r.get('nota_final') is not None else '-'),
                ])
        elements.append(_pdf_table(data))
        return _pdf_response(elements, 'reporte_calificaciones.pdf', landscape_mode=True)

    wb = Workbook()
    ws = wb.active
    ws.title = "Calificaciones"
    if id_materia and id_seccion and id_periodo and not id_estudiante:
        headers = ['Cédula', 'Nombres', 'Apellidos', 'N1', 'N2', 'N3', 'Examen Final', 'Promedio']
        ws.append(headers)
        for r in rows:
            ws.append([
                r.get('cedula'), r.get('nombres'), r.get('apellidos'),
                r.get('nota1'), r.get('nota2'), r.get('nota3'),
                r.get('examen_final'), r.get('nota_final')
            ])
    else:
        headers = ['Estudiante', 'Cédula', 'Materia', 'Periodo', 'N1', 'N2', 'N3', 'Examen', 'Promedio']
        ws.append(headers)
        for r in rows:
            ws.append([
                r.get('estudiante_nombre') or f"{r.get('nombres', '')} {r.get('apellidos', '')}",
                r.get('cedula'), r.get('materia_nombre'), r.get('periodo_nombre'),
                r.get('nota1'), r.get('nota2'), r.get('nota3'),
                r.get('examen_final'), r.get('nota_final')
            ])
    hf, hfill = _header_style()
    for col in range(1, len(headers) + 1):
        ws.cell(1, col).font = hf
        ws.cell(1, col).fill = hfill
    return _excel_response(wb, 'reporte_calificaciones.xlsx')